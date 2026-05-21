import json
import sys

from openpyxl import load_workbook


def main() -> None:
    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    year = worksheet["C7"].value
    rows = []

    for row in worksheet.iter_rows(min_row=8, values_only=True):
        country = row[1]
        value = row[2]

        if not country or not isinstance(value, (int, float)):
            continue

        if isinstance(country, str) and "human development" in country.lower():
            continue

        rows.append(
            {
                "country": country,
                "value": float(value),
                "year": int(year),
                "source": "UNDP Human Development Report 2025",
            }
        )

    print(json.dumps(rows))


if __name__ == "__main__":
    main()
